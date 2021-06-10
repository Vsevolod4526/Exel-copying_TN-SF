from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
import time
import datetime
import os
from tkinter import messagebox
import openpyxl
import copy
import xlrd



root=Tk()
root.geometry('700x820+500+20')

root.title("Копирование данных из отчётов .xls")


f_top = Frame()
b_spravka=Button(f_top,text='Справка о правилах использования программы',width=85,height=1,fg='black')


b1=Button(text='Выбрать директорию с файлами',width=85,height=1,bg="snow3")
l1=Label(text='Укажите путь в папку с товарными накладными',font='Arial 16')
l2=Label(text=' с расширением .xlsx',font='Arial 16')
e1=Entry(width=100)

b2=Button(text='Запуск копирования',width=85,height=1,bg='cornflower blue',fg='white')
l2marker=Label(text='данные ТН пока не скопированы ',font='Arial 10')


e1promegTN=Entry(width=100)
b1promeg=Button(text='Выбрать директорию для сохранения промежуточного массива из ТН в EXEL',width=85,height=1)
b2promeg=Button(text='Запуск Сохранения',width=85,height=1)





l4=Label(text='     ',font='Arial 5')
l_1Val=Label(text='Выбор файла с данными по валюте     ',font='Arial 15')
e_1Val=Entry(width=100)
b_2Val=Button(text='1 - Выбрать путь к файлу ',width=85,height=1)
b_3Val=Button(text='2 - Сохранить данные по валюте ',width=85,height=1)
l_2Val=Label(text='Данные из файла с валютой пока не скопированы',font='Arial 10')
b_4Val=Button(text='3 - Интегрировать данные валюты ',width=85,height=1)
l_3Val=Label(text='Данные валют пока не интегрированы ',font='Arial 10')



l21=Label(text='Укажите путь в папку со счетами фактуры',font='Arial 16')
l22=Label(text=' с расширением .xls или xlsx',font='Arial 16')
e21=Entry(width=100)
b21=Button(text='Выбрать директорию с файлами',width=85,height=1,bg="snow3")
b22=Button(text='Запуск копирования',width=85,height=1,bg='cornflower blue',fg='white')
l22marker=Label(text='данные CФ пока не скопированы ',font='Arial 10')
e21promegSF=Entry(width=100)
b21promeg=Button(text='Выбрать директорию для сохранения промежуточного массива из CФ в EXEL',width=85,height=1)
b22promeg=Button(text='Запуск Сохранения',width=85,height=1)

l23=Label(text='  ',font='Arial 5')


l5=Label(text='Укажите путь для создания стокового EXEL',font='Arial 16')
e5=Entry(width=100)
b5=Button(text='Выбор директории для стокового EXEL',width=85,height=1,bg="snow3")
b52=Button(text='Запуск создания стокового EXEL',width=85,height=1,fg='white',bg='royalblue3')




f_top.pack()
b_spravka.pack(side=RIGHT)


l1.pack()
l2.pack()
e1.pack()
b1.pack()
b2.pack()
l2marker.pack()
e1promegTN.pack()
b1promeg.pack()
b2promeg.pack()

l4.pack()







l21.pack()
l22.pack()
e21.pack()
b21.pack()
b22.pack()
l22marker.pack()
e21promegSF.pack()
b21promeg.pack()
b22promeg.pack()
l23.pack()

l_1Val.pack()
e_1Val.pack()
b_2Val.pack()
b_3Val.pack()
l_2Val.pack()
b_4Val.pack()
l_3Val.pack()

l5.pack()
e5.pack()
b5.pack()
b52.pack()






def show_spravka():
    messagebox.showinfo(" Для корректной работы программы важно : ", ("1. Чтобы все файлы,которые используются в программt были закрыты(файлы ТН, файлы СФ,файл с валютой) \n"+
                                                                    "2. Чтобы в папке,которую пользователь выбирает для копирования ТН,находились ТОЛЬКО файлы ТН.То же верно и для выбора папки с файлами СФ \n "+
                                                                    "3. Выполнять интеграцию в файлы валюты и создавать стоковый exel только после того,как данные ТН и СФ были скопированы \n"+
                                                                    " Дополнительно : Выбирать директорию для сохранения промежуточного массива и сохранять его -  не обязательно, это функция опциональна \n"+
                                                                    "Общие рекомендации : нажимать кнопки по порядку сверху вниз "))
                        
    return


def ask_Val_xl():
    askdir=filedialog.askopenfilename()
    e_1Val.delete(0, last=END)
    e_1Val.insert(0,askdir)
    print(askdir)


def save_from_Val():
    if 'spisok_massivov_dla_zapici' in globals():
        if len(spisok_massivov_dla_zapici)>0:
            pass
        else:
            messagebox.showinfo("Отсутствие данных", "Данные из ТН не были загружены")
            return
    else:
        messagebox.showinfo("Отсутствие данных", "Данные из ТН не были загружены")
        return


    try:

        wb_VAL_xlrd=xlrd.open_workbook(filename=e_1Val.get())
        sheet_xlrd = wb_VAL_xlrd.sheet_by_index(0)

        kol_row=sheet_xlrd.nrows
        kol_column=sheet_xlrd.ncols
        global massiv_s_valutoy
        massiv_s_valutoy=[]
        for i in range(sheet_xlrd.nrows):
            tmp=[]
            tmp.append(sheet_xlrd.cell(i,0).value)
            tmp.append(sheet_xlrd.cell(i,1).value)
            tmp.append(sheet_xlrd.cell(i,2).value)
            massiv_s_valutoy.append(tmp)



        
        print('Данные из файла с валютой успешно сохранены')
    except:
        messagebox.showinfo("Ошибка копирования ", "Ошибка в копировании данных из предполагаемого файла с  валютой")
        return
        
    l_2Val['bg']="#007dff"
    l_2Val['text']='Данные из файла с валютами скопированы'


def integration_Val_v_TN():
    if 'spisok_massivov_dla_zapici' in globals() and 'massiv_s_valutoy' in globals():
        if len(spisok_massivov_dla_zapici)>0 and len(massiv_s_valutoy)>0:
            pass
        else:
            messagebox.showinfo("Отсутствие данных", "Данные из ТН и\или файла с валютой не были загружены")
            return
    else:
        messagebox.showinfo("Отсутствие данных", "Данные из ТН и\или файла с валютой не были загружены")
        return

    try:
        for i in spisok_massivov_dla_zapici:
            for j in massiv_s_valutoy:
                if i[15]==j[2]: #поиск совпадений по дате и коду валюты and i[17]==int(j[1])
                    
                    if i[17]==840:#если есть цена в долларах
                        if int(j[1])==840:
                            i[11]=j[0]
                            if i[6]!='-' and i[6]!=0: 
                                i[7]=round(i[6]*i[11],2)
                    if i[17]==978:#если есть цена в евро
                        if int(j[1])==978:
                            i[10]=j[0]
                            if i[5]!='-' and i[5]!=0: 
                                i[7]=round(i[5]*i[10],2)
                    if i[17]==643:# если цена в рублях
                        if int(j[1])==978:#  а нашло евро
                            i[10]=j[0]
                            if i[7]!='-' and i[7]!=0:
                                i[5]=round(i[7]/i[10],2)
                        if int(j[1])==840:#нашло в долларах
                            i[11]=j[0]
                            if i[7]!='-' and i[7]!=0:
                                i[6]=round(i[7]/i[11],2)
                    
            for k in massiv_s_valutoy:
                if i[15]==k[2] and i[7]!='-':#доллары
                    if int(k[1])==978:#  а нашло евро
                            i[10]=k[0]
                            if i[7]!='-' and i[7]!=0:
                                i[5]=round(i[7]/i[10],2)
                    
                    
                if i[15]==k[2] and i[7]!='-':#евро
                    if int(k[1])==840:#нашло в долларах
                            i[11]=k[0]
                            if i[7]!='-' and i[7]!=0:
                                i[6]=round(i[7]/i[11],2)

    except:
        messagebox.showinfo("Ошибка при интеграции данных", "Ошибка при интеграции данных о валюте ")
        return

        
                
                
                
                                
                        
                    
                    
                    
    l_3Val['bg']="#007dff"
    l_3Val['text']='Данные валют интегрированы'

       # 1 счет-фактура №
        # 2 дата
        # 3 наименование из ТН (часть)
        # 4 вес
        # 5 стоимость в евро
        # 6 стоимость в долларах
        # 7 стоимость в рублях
        # 8 страна
        # 9 номер дт
        # 10 курс евро
        # 11 курс доллара
        # 12 кол-во
        # 13 код
        # 14 ТН номер
        # 15 дата
        # 16 счет на оплату
        # 17 КОД ВАЛЮТЫ
        # 18 код товара из СФ
        # 19 код товара из ТН
        # 20
                
    








def askdir():
    askdir=filedialog.askdirectory()
    e1.delete(0, last=END)
    e1.insert(0,askdir)
    print(askdir)


def askdir_promeg_TN():
    askdir=filedialog.askdirectory()
    e1promegTN.delete(0, last=END)
    e1promegTN.insert(0,askdir)
    print(askdir)

def promeg_TN_save():
    if 'spisok_massivov_dla_zapici' in globals():
        if len(spisok_massivov_dla_zapici)>0:
            pass
        else:
            messagebox.showinfo("Отсутствие данных", "Данные из ТН не были загружены")
            return    
    else:
        messagebox.showinfo("Отсутствие данных", "Данные из ТН не были загружены")
        return
    
    try:
        os.chdir(e1promegTN.get())
    except OSError:
        messagebox.showerror("Ошибка","Ошибка в пути папки. Директория не найдена")
        return

    stock_EXEL_fail_workbook=openpyxl.Workbook()
    sheet_of_stock_EXEL = stock_EXEL_fail_workbook[stock_EXEL_fail_workbook.sheetnames[0]]

    sheet_of_stock_EXEL.cell(row = 1, column = 1).value="Счет-фактура №"
    sheet_of_stock_EXEL.cell(row = 1, column = 2).value="Дата СФ"
    sheet_of_stock_EXEL.cell(row = 1, column = 3).value="Наименование"
    sheet_of_stock_EXEL.cell(row = 1, column = 4).value="Вес"
    sheet_of_stock_EXEL.cell(row = 1, column = 5).value="Стоимость в евро"
    sheet_of_stock_EXEL.cell(row = 1, column = 6).value="Стоимость в долларах"
    sheet_of_stock_EXEL.cell(row = 1, column = 7).value="Стоимость в рублях"
    sheet_of_stock_EXEL.cell(row = 1, column = 8).value="Страна"
    sheet_of_stock_EXEL.cell(row = 1, column = 9).value="Номер дт"
    sheet_of_stock_EXEL.cell(row = 1, column = 10).value="Курс евро"
    sheet_of_stock_EXEL.cell(row = 1, column = 11).value="Курс доллара"
    sheet_of_stock_EXEL.cell(row = 1, column = 12).value="Кол-во"
    sheet_of_stock_EXEL.cell(row = 1, column = 13).value="Код"
    sheet_of_stock_EXEL.cell(row = 1, column = 14).value="Товарная Накладная"# ТН - это товарная накладная????
    sheet_of_stock_EXEL.cell(row = 1, column = 15).value="Дата ТН"
    sheet_of_stock_EXEL.cell(row = 1, column = 16).value="Счет на оплату"
    sheet_of_stock_EXEL.cell(row = 1, column = 17).value="Код Валюты"
    sheet_of_stock_EXEL.cell(row = 1, column = 18).value="СФ Код товара/ работ, услуг"
    sheet_of_stock_EXEL.cell(row = 1, column = 19).value="ТН Код товара/ работ, услуг"
    sheet_of_stock_EXEL.cell(row = 1, column = 20).value=""
    sheet_of_stock_EXEL.cell(row = 1, column = 21).value="Порядковый номер СФ"
    sheet_of_stock_EXEL.cell(row = 1, column = 22).value="Порядковый номер ТН"
    
    list_time=time.localtime()
    beta_name=" TN Стоковый_EXEL "+str(list_time[2])+'.'+str(list_time[1])+'.'+str(list_time[0])+'  '+str(list_time[3])+'-'+str(list_time[4])+".xlsx"
    beta_name=list(beta_name)
    name=''
    for i in range(len(beta_name)):
        if beta_name[i]==":":
            beta_name[i]="-"
    for i in beta_name:
        name+=i
     #ЗАПИСЬ ГЛОБАЛЬНОГО МАССИВА ИЗ ТН
    try:
        for a in range(len(spisok_massivov_dla_zapici)):
            #spisok_massivov_dla_zapici[a]
            for d in range(1,23):
                sheet_of_stock_EXEL.cell(row = a+2, column = d).value=spisok_massivov_dla_zapici[a][d]
        stock_EXEL_fail_workbook.save(name)
    except:
        messagebox.showerror("Ошибка","Ошибка в записи данных в EXEL")
        print("Ошибка записи файлов из ТН в EXEL")
        


    
    print("Создание стокового EXEL завершено")












    
    
    
    


def askdir_of_SF():
    askdir=filedialog.askdirectory()
    e21.delete(0, last=END)
    e21.insert(0,askdir)
    print(askdir)

def askdir_promeg_SF():
    askdir=filedialog.askdirectory()
    e21promegSF.delete(0, last=END)
    e21promegSF.insert(0,askdir)
    print(askdir)

def promeg_SF_save():
    if 'spisok_massivov_is_SF' in globals(): ###
        if len(spisok_massivov_is_SF)>0:
            pass
        else:
            messagebox.showinfo("Отсутствие данных", "Данные из СФ не были загружены")
            return    
    else:
        messagebox.showinfo("Отсутствие данных", "Данные из СФ не были загружены")
        return
    
    try:
        os.chdir(e21promegSF.get())
    except OSError:
        messagebox.showerror("Ошибка","Ошибка в пути папки. Директория не найдена")
        return

    stock_EXEL_fail_workbook=openpyxl.Workbook()
    sheet_of_stock_EXEL = stock_EXEL_fail_workbook[stock_EXEL_fail_workbook.sheetnames[0]]

    sheet_of_stock_EXEL.cell(row = 1, column = 1).value="Счет-фактура №"
    sheet_of_stock_EXEL.cell(row = 1, column = 2).value="Дата СФ"
    sheet_of_stock_EXEL.cell(row = 1, column = 3).value="Наименование"
    sheet_of_stock_EXEL.cell(row = 1, column = 4).value="Вес"
    sheet_of_stock_EXEL.cell(row = 1, column = 5).value="Стоимость в евро"
    sheet_of_stock_EXEL.cell(row = 1, column = 6).value="Стоимость в долларах"
    sheet_of_stock_EXEL.cell(row = 1, column = 7).value="Стоимость в рублях"
    sheet_of_stock_EXEL.cell(row = 1, column = 8).value="Страна"
    sheet_of_stock_EXEL.cell(row = 1, column = 9).value="Номер дт"
    sheet_of_stock_EXEL.cell(row = 1, column = 10).value="Курс евро"
    sheet_of_stock_EXEL.cell(row = 1, column = 11).value="Курс доллара"
    sheet_of_stock_EXEL.cell(row = 1, column = 12).value="Кол-во"
    sheet_of_stock_EXEL.cell(row = 1, column = 13).value="Код"
    sheet_of_stock_EXEL.cell(row = 1, column = 14).value="Товарная Накладная"# ТН - это товарная накладная????
    sheet_of_stock_EXEL.cell(row = 1, column = 15).value="Дата ТН"
    sheet_of_stock_EXEL.cell(row = 1, column = 16).value="Счет на оплату"
    sheet_of_stock_EXEL.cell(row = 1, column = 17).value="Код Валюты"
    sheet_of_stock_EXEL.cell(row = 1, column = 18).value="СФ Код товара/ работ, услуг"
    sheet_of_stock_EXEL.cell(row = 1, column = 19).value="ТН Код товара/ работ, услуг"
    sheet_of_stock_EXEL.cell(row = 1, column = 20).value=""
    sheet_of_stock_EXEL.cell(row = 1, column = 21).value="Порядковый номер СФ"
    sheet_of_stock_EXEL.cell(row = 1, column = 22).value="Порядковый номер ТН"

    list_time=time.localtime()
    beta_name=" СФ Стоковый_EXEL "+str(list_time[2])+'.'+str(list_time[1])+'.'+str(list_time[0])+'  '+str(list_time[3])+'-'+str(list_time[4])+".xlsx"
    beta_name=list(beta_name)
    name=''
    for i in range(len(beta_name)):
        if beta_name[i]==":":
            beta_name[i]="-"
    for i in beta_name:
        name+=i
     #ЗАПИСЬ ГЛОБАЛЬНОГО МАССИВА ИЗ ТН
    try:
        for a in range(len(spisok_massivov_is_SF)):
            #spisok_massivov_dla_zapici[a]
            for d in range(1,23):
                sheet_of_stock_EXEL.cell(row = a+2, column = d).value=spisok_massivov_is_SF[a][d]
        stock_EXEL_fail_workbook.save(name)
    except:
        messagebox.showerror("Ошибка","Ошибка в записи данных в EXEL")
        print("Ошибка записи файлов из ТН в EXEL")
        


    
    print("Создание стокового EXEL завершено")





























def askdir_of_stock():
    askdir=filedialog.askdirectory()
    e5.delete(0, last=END)
    e5.insert(0,askdir)
    print(askdir_of_stock)






def zapusk_kopy_SF():

    global spisok_massivov_is_SF

    spisok_massivov_is_SF=[]

    try:
        os.chdir(e21.get())
    except OSError:
        messagebox.showerror("Ошибка","Ошибка в пути папки. Директории не существует")

    list_of_SF=os.listdir(os.getcwd())
    
    list_of_SF_xl=[]
    for i in range(len(list_of_SF)):
        if '.xls' in list_of_SF[i] or '.xlsx' in list_of_SF[i] or '.XLSX' in list_of_SF[i] or '.XLS' in list_of_SF[i]:
            list_of_SF_xl.append(list_of_SF[i])

    #print(list_of_SF_xl)
    
    


    #начало открывания  экселей из списка
    for l in range(len(list_of_SF_xl)):
        
        

        wb_sf_xlrd=xlrd.open_workbook(filename=list_of_SF_xl[l])
        sheet_xlrd = wb_sf_xlrd.sheet_by_index(0)

        kol_row=sheet_xlrd.nrows
        kol_column=sheet_xlrd.ncols

        
        
        nomer_SF='-'
        data_SF='-'
        # поиск номера СФ и Даты
        succes=0
        try:
            for rown in  range(0,kol_row):
                
                for coln in range(0,kol_column):
                    
                    if sheet_xlrd.cell(rown,coln).value!=None:
                        
                        if type(sheet_xlrd.cell(rown,coln).value)==str:
                            if "счет-фактура n" in sheet_xlrd.cell(rown,coln).value.lower():
                                #print(sheet_xlrd.cell(rown,coln).value,' поле счет-фактура n  найдено')
                                for icol in range(1,25):
                                    if ('1' in sheet_xlrd.cell(rown,coln+icol).value or
                                        '2' in sheet_xlrd.cell(rown,coln+icol).value or
                                        '3' in sheet_xlrd.cell(rown,coln+icol).value or
                                        '4' in sheet_xlrd.cell(rown,coln+icol).value or
                                        '5' in sheet_xlrd.cell(rown,coln+icol).value or
                                        '6' in sheet_xlrd.cell(rown,coln+icol).value or
                                        '7' in sheet_xlrd.cell(rown,coln+icol).value or
                                        '8' in sheet_xlrd.cell(rown,coln+icol).value or
                                        '9' in sheet_xlrd.cell(rown,coln+icol).value):
                                        
                                        nomer_SF=sheet_xlrd.cell(rown,coln+icol).value
                                        #print(nomer_SF)

                                        for jcol in range(1,30):
                                            if ((sheet_xlrd.cell(rown,coln+icol+jcol).value!=None and len(sheet_xlrd.cell(rown,coln+icol+jcol).value)>4) or
                                            type(sheet_xlrd.cell(rown,coln+icol+jcol).value)==type(datetime.datetime.today())):
                                            
                                                #print(sheet_xlrd.cell(rown,coln+icol+jcol).value)
                                                data_SF=sheet_xlrd.cell(rown,coln+icol+jcol).value
                                                succes=1
                                                raise Exception
                                            
        except:
            #print("поля найдены,выход из циклов")
            pass
            
                                                
        if succes!=1:
            messagebox.showerror("Ошибка","Ошибка, в файле "+str(list_of_SF_xl[l])+" не были найдены данные")
            succes=0


        #начало поиска клетки для подсчета строк и нужных столбцов



        row_pole_Npp=0
        column_pole_Npp=0

        schetchik_strok=0
        sposok_nomerov_strok_dla_copirovania=[]

        row_pole_A=0
        column_pole_A=0
        
        
        try:
            
            for rown in  range(5,40):
                for coln in range(0,10):
                    if sheet_xlrd.cell(rown,coln).value!=None:
                        if type(sheet_xlrd.cell(rown,coln).value)==str:   
                            if "n п/п" in sheet_xlrd.cell(rown,coln).value.lower():
                                
                                #print(sheet_xlrd.cell(rown,coln).value)
                                row_pole_Npp=rown
                                column_pole_Npp=coln
                                for jrow in range(1,10):
                                    if "а" in sheet_xlrd.cell(rown+jrow,coln).value.lower() or "a" in sheet_xlrd.cell(rown+jrow,coln).value.lower():
                                        #print(sheet_xlrd.cell(rown+jrow,coln).value)
                                        row_pole_A=rown+jrow
                                        column_pole_A=coln

                                        for vcol in range(row_pole_A+1,kol_row): # vcol перепутано с vrow (правильно наоборот в плане названия)  
                                            if type(sheet_xlrd.cell(vcol,coln).value)==int:
                                                #print(sheet_xlrd.cell(vcol,coln).value)
                                                sposok_nomerov_strok_dla_copirovania.append(vcol)
                                            if type(sheet_xlrd.cell(vcol,coln).value)==str:
                                                if sheet_xlrd.cell(vcol,coln).value.replace(' ','').isdigit()==True:
                                                    #print(sheet_xlrd.cell(vcol,coln).value)
                                                    sposok_nomerov_strok_dla_copirovania.append(vcol)
                                            if type(sheet_xlrd.cell(vcol,coln).value)==str:
                                                if sheet_xlrd.cell(vcol,coln).value.replace(' ','').isalpha():
                                                    succes=1
                                                    raise Exception
                                                
                                          
                                                
                                                #print(type(sheet_xlrd.cell(vcol,coln).value),' ',sheet_xlrd.cell(vcol,coln).value)
                                        
                                
                                
                                
                            
                                    
                                    
                                        
        except:
            pass
            
            
        if succes!=1:
            messagebox.showerror("Ошибка","Ошибка, в файле "+str(list_of_SF_xl[l])+" не были найдены данные по столбцам и строкам")
            succes=0
        

    

        #поиск столбцов и формирование списков
        list_of_kod_tovar_uslug=[]
        list_of_kod_vida_tovara=[]
        list_of_nomer_tamojnoy_declaracii=[]
        list_of_strana=[]
        try:
            for icolumn in range(column_pole_Npp,kol_column):
                if ("код товара" in sheet_xlrd.cell(row_pole_Npp,icolumn).value.lower() and
                    "работ" in sheet_xlrd.cell(row_pole_Npp,icolumn).value and
                    "услуг" in sheet_xlrd.cell(row_pole_Npp,icolumn).value):
                    #print(sheet_xlrd.cell(row_pole_Npp,icolumn).value)
                    for nom_row in sposok_nomerov_strok_dla_copirovania:
                        list_of_kod_tovar_uslug.append(sheet_xlrd.cell(nom_row,icolumn).value)
                if "код вида товара" in sheet_xlrd.cell(row_pole_Npp,icolumn).value.lower():
                    #print(sheet_xlrd.cell(row_pole_Npp,icolumn).value)
                    for nom_row in sposok_nomerov_strok_dla_copirovania:
                        list_of_kod_vida_tovara.append(sheet_xlrd.cell(nom_row,icolumn).value)

                        #номер таможенной декларации
                if "номер таможенной декларации" in sheet_xlrd.cell(row_pole_Npp,icolumn).value.lower():
                    #print(sheet_xlrd.cell(row_pole_Npp,icolumn).value)
                    for nom_row in sposok_nomerov_strok_dla_copirovania:
                        if len(sheet_xlrd.cell(nom_row,icolumn).value)>5:
                            nomer_DT=sheet_xlrd.cell(nom_row,icolumn).value+sheet_xlrd.cell(nom_row+1,icolumn).value
                            list_of_nomer_tamojnoy_declaracii.append(nomer_DT)
                        else:
                            if sheet_xlrd.cell(nom_row,icolumn).value!=None:
                                list_of_nomer_tamojnoy_declaracii.append(sheet_xlrd.cell(nom_row,icolumn).value)
                            
                if ('10a' in sheet_xlrd.cell(row_pole_A,icolumn).value.lower() or
                    '10а' in sheet_xlrd.cell(row_pole_A,icolumn).value.lower()):
                    #print(sheet_xlrd.cell(row_pole_A,icolumn).value)
                    for nom_row in sposok_nomerov_strok_dla_copirovania:
                        strana=''
                        for b in sheet_xlrd.cell(nom_row,icolumn).value:
                            if b!='(' and b.isalpha()==True:
                                strana+=b
                                
                            else:
                                break
                        list_of_strana.append(strana)
        except:
            print(messagebox.showerror("Ошибка","Ошибка, в файле "+str(list_of_SF_xl[l])+" не были найдены данные по столбцам и строкам(2)"))
        
                        
                
            
        
        #Поиск итоговой суммы без НДС

        itogovaya_summa_bez_NDS_is_SF=0
        #sposok_nomerov_strok_dla_copirovania

        #Поиск колонки с суммой без НДС
            

        n_column_itogovoy_summy_bez_NDS=0

        for i_col in range(3,kol_column):
            if sheet_xlrd.cell(row_pole_Npp,i_col).value!=None:
                #print(sheet_xlrd.cell(row_pole_Npp,i_col).value)
                if type(sheet_xlrd.cell(row_pole_Npp,i_col).value)==str:
                    if ("стоимость товаров" in sheet_xlrd.cell(row_pole_Npp,i_col).value.lower()
                    and "без налога" in sheet_xlrd.cell(row_pole_Npp,i_col).value.lower()):
                        n_column_itogovoy_summy_bez_NDS=i_col
                        break
        n_row_itogovoy_summy_bez_NDS=0
        # Вычисление строки поля "Всего к оплате"
        for n_row in range(sposok_nomerov_strok_dla_copirovania[-1],kol_row):
            for n_col in range(1,kol_column):
                if sheet_xlrd.cell(n_row,n_col).value!=None:
                    if type(sheet_xlrd.cell(n_row,n_col).value)==str:
                        if "всего к оплате" == sheet_xlrd.cell(n_row,n_col).value.lower().strip():
                            #print(sheet_xlrd.cell(n_row,n_col).value)
                            n_row_itogovoy_summy_bez_NDS=n_row
                            break
                            
                        
                    
        itogovaya_summa_bez_NDS_is_SF=sheet_xlrd.cell(n_row_itogovoy_summy_bez_NDS,n_column_itogovoy_summy_bez_NDS).value
        
        print()
                        
            
                                
                                
            
        

        

        #Формирование массива из списка

        #формирование списка для записи в exel
        # 1 счет-фактура №
        # 2 дата
        # 3 наименование
        # 4 вес
        # 5 стоимость в евро
        # 6 стоимость в долларах
        # 7 стоимость в рублях
        # 8 страна
        # 9 номер дт
        # 10 курс евро
        # 11 курс доллара
        # 12 кол-во
        # 13 код
        # 14 ТН номер
        # 15 дата
        # 16 счет на оплату
        # 17 наименование товара из ТН
        # 18 код товара из СФ
        # 19 код товара из ТН
        #
        # 21 Порядковый номер товара из СФ
        # 22 Порядковый номер товара из ТН
        # 23 Итоговая сумма без НДС из ТН
        # 24 Итоговая сумма без НДС из СФ

        for index in range(len(sposok_nomerov_strok_dla_copirovania)):
            row_list=['-' for jk in range(27)]
            row_list[1]=nomer_SF
            row_list[2]=data_SF
            row_list[8]=list_of_strana[index]
            row_list[9]=list_of_nomer_tamojnoy_declaracii[index]
            row_list[13]=list_of_kod_vida_tovara[index]
            row_list[18]=list_of_kod_tovar_uslug[index]
            row_list[21]=index+1
            row_list[24]=round(itogovaya_summa_bez_NDS_is_SF,1)
            

            spisok_massivov_is_SF.append(row_list)

            print(row_list)
        
    l22marker['bg']="#007dff"
    l22marker['text']='Данные СФ скопированы'

            
        

    
    

def zapusk_kopy():

    global spisok_massivov_dla_zapici
    

    spisok_massivov_dla_zapici=[]
    try:
        os.chdir(e1.get())
    except OSError:
        messagebox.showerror("Ошибка","Ошибка в пути папки. Директории не существует")

    list_of_tovarni_nakladniyi=os.listdir(os.getcwd())
    list_of_tovarni_nakladniyi_xl=[]
    for i in range(len(list_of_tovarni_nakladniyi)):
        if '.xls' in list_of_tovarni_nakladniyi[i] or '.xlsx' in list_of_tovarni_nakladniyi[i] or '.XLSX' in list_of_tovarni_nakladniyi[i] or '.XLS' in list_of_tovarni_nakladniyi[i]:
            list_of_tovarni_nakladniyi_xl.append(list_of_tovarni_nakladniyi[i])

    print(list_of_tovarni_nakladniyi_xl)

    #начало открывания списка экселей
    for l in range(len(list_of_tovarni_nakladniyi_xl)):
        
        wb=openpyxl.load_workbook(filename =list_of_tovarni_nakladniyi_xl[l])
        list_of_sheets=wb.sheetnames
        sheet_0=wb[list_of_sheets[0]]


        kol_row=0
        kol_column=0
        nomer_documenta=0
        Data_sostavleniya=''
        succes=0

        for row in sheet_0.iter_rows():
            kol_column=len(row)
            
            break

        for column in sheet_0.iter_cols():
            kol_row=len(column)
            break

        try:
            for rown in  range(1,kol_row):
                for coln in range(1,kol_column):
                    if sheet_0.cell(row=rown,column=coln).value!=None:
                        if type(sheet_0.cell(row=rown,column=coln).value)==str:
                            if "Номер документа" in sheet_0.cell(row=rown,column=coln).value:
                                
                                

                                for i in range(1,5):
                                    if type(sheet_0.cell(row=rown+i,column=coln).value)==int:
                                        #print("Ячейка с предполагаемым номером документа найдена")
                                        nomer_documenta=sheet_0.cell(row=rown+i,column=coln).value
                                        #print(nomer_documenta)

                                        for j in range(1,10):
                                            if sheet_0.cell(row=rown+i,column=coln+j).value!=None:
                                                #print("Ячейка с предполагаемой датой найдена")
                                                #print(sheet_0.cell(row=rown+i,column=coln+j).value)
                                                if type(sheet_0.cell(row=rown+i,column=coln+j).value)==type(datetime.datetime.now()):
                                                    if len(str(sheet_0.cell(row=rown+i,column=coln+j).value.day))==1:
                                                        day='0'+str(sheet_0.cell(row=rown+i,column=coln+j).value.day)
                                                    else:
                                                        day=str(sheet_0.cell(row=rown+i,column=coln+j).value.day)
                                                    if len(str(sheet_0.cell(row=rown+i,column=coln+j).value.month))==1:
                                                        month='0'+str(sheet_0.cell(row=rown+i,column=coln+j).value.month)
                                                    else:
                                                        month=str(sheet_0.cell(row=rown+i,column=coln+j).value.month)
                                                        
                                                    Data_sostavleniya=day+'.'+month+'.'+str(sheet_0.cell(row=rown+i,column=coln+j).value.year)
                                                else:
                                                    Data_sostavleniya=str(sheet_0.cell(row=rown+i,column=coln+j).value)
                                                #print(str(Data_sostavleniya))#МОЖНО ПЕРЕВЕСТИ В СТРОКУ !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
                                                succes=1
                                                raise Exception

        except:
            #print("поля найдены,выход из циклов")
            pass
            
                                                
        if succes!=1:
            messagebox.showerror("Ошибка","Ошибка, в файле "+str(list_of_tovarni_nakladniyi_xl[l])+" не были найдены данные")

        #Поиск поля "Номер по порядку" и выяснение количества строк для копирования
        row_pole_NOMER_PO_PORADKY=1
        column_pole_NOMER_PO_PORADKY=1
        kol_strok_dla_copirovania=0
        try:
            
            for rown in  range(19,27):
                for coln in range(1,8):
                    
                    if sheet_0.cell(row=rown,column=coln).value!=None:
                        if type(sheet_0.cell(row=rown,column=coln).value)==str:
                            if "Номер по порядку" in sheet_0.cell(row=rown,column=coln).value:
                                #print(sheet_0.cell(row=rown,column=coln).value)
                                row_pole_NOMER_PO_PORADKY=rown
                                column_pole_NOMER_PO_PORADKY=coln
                                    
                                    
                                        
        except:
            #print(kol_strok_dla_copirovania)
            print()
        flag=0
        rows_list_to_find=[]
        for i in range(row_pole_NOMER_PO_PORADKY,kol_row):
            #print(sheet_0.cell(row=i,column=column_pole_NOMER_PO_PORADKY).value)
            if sheet_0.cell(row=i,column=column_pole_NOMER_PO_PORADKY).value!=None:
                if type(sheet_0.cell(row=i,column=column_pole_NOMER_PO_PORADKY).value)==int:
                    if sheet_0.cell(row=i,column=column_pole_NOMER_PO_PORADKY).value==1 and flag==0:
                        flag=1
                        continue
                    if sheet_0.cell(row=i,column=column_pole_NOMER_PO_PORADKY).value==kol_strok_dla_copirovania+1:
                        kol_strok_dla_copirovania+=1
                        rows_list_to_find.append(i)
                
                    

        
        #print("KOL-VO STROK ", kol_strok_dla_copirovania)
        #print("Stroki dla copirovania ",rows_list_to_find)





        
        schet_na_oplaty=''
        try:
            for nrow in  range(6,row_pole_NOMER_PO_PORADKY):
                for ncol in range(3,kol_column):
                    
                    if sheet_0.cell(row=nrow,column=ncol).value!=None:
                        if type(sheet_0.cell(row=nrow,column=ncol).value)==str:
                            if 'Счет на оплату' in sheet_0.cell(row=nrow,column=ncol).value:
                                #print(sheet_0.cell(row=nrow,column=ncol).value)
                                n_y=sheet_0.cell(row=nrow,column=ncol).value.find('у')
                                pole=sheet_0.cell(row=nrow,column=ncol).value
                                for q in range(n_y+2,len(pole)):
                                    if pole[q]!=' ':
                                        schet_na_oplaty+=pole[q]
                                    else:
                                        raise Exception
                                    
        except:
            pass

        


        currency=''
        #поиск номеров нужных для копирования столбцов
        # вес , стоимость ,количество
        list_of_prices=[]
        list_of_kol_and_massa=[]
        list_of_naimenovaniy_tovar=[]
        column_stolbca_s_summoy_bez_nds=0 # ПЕРЕМЕННАЯ С НОМЕРОМ СТОЛБЦА ДЛЯ ПОИСКА ИТОГОВОЙ СУММЫ
        for i in range(1,kol_column):
            if sheet_0.cell(row=row_pole_NOMER_PO_PORADKY,column=i).value!=None:
                if type(sheet_0.cell(row=row_pole_NOMER_PO_PORADKY,column=i).value)==str:
                    kletka=sheet_0.cell(row=row_pole_NOMER_PO_PORADKY,column=i).value.lower()
                    
                
                    if "(масса нетто)" in kletka:
                        for nrow in rows_list_to_find:
                            list_of_kol_and_massa.append(sheet_0.cell(row=nrow,column=i).value)
                            
                        
                    if "Цена" in sheet_0.cell(row=row_pole_NOMER_PO_PORADKY,column=i).value:
                        kletka_s_kol_deneg=sheet_0.cell(row=row_pole_NOMER_PO_PORADKY+1,column=i).value.lower()
                        if "евр" in kletka_s_kol_deneg:
                            currency='Евро'
                            #print(currency)
                        if "дол" in kletka_s_kol_deneg:
                            currency='Дол'
                            #print(currency)
                        if "руб" in kletka_s_kol_deneg:
                            currency='Руб'
                            #print(currency)
                        #for nrow in rows_list_to_find:
                            #print(sheet_0.cell(row=nrow,column=i).value)
                            #list_of_prices.append(sheet_0.cell(row=nrow,column=i).value)# удалено, исправление к другой колонке
                    if "товар" in sheet_0.cell(row=row_pole_NOMER_PO_PORADKY,column=i).value.lower():
                        for nrow in rows_list_to_find:
                            list_of_naimenovaniy_tovar.append(sheet_0.cell(row=nrow,column=i).value)
                    if "сумма без учета" in sheet_0.cell(row=row_pole_NOMER_PO_PORADKY,column=i).value.lower():
                        column_stolbca_s_summoy_bez_nds=i
                        for nrow in rows_list_to_find:
                            list_of_prices.append(sheet_0.cell(row=nrow,column=i).value)

        #Поиск кода из ТН
        spisok_kodov_is_TN=[]
        succes=0
        try:
            for inrow in range(row_pole_NOMER_PO_PORADKY,row_pole_NOMER_PO_PORADKY+15):
                if (sheet_0.cell(row=inrow,column=column_pole_NOMER_PO_PORADKY).value==1 or
                    (type(sheet_0.cell(row=inrow,column=column_pole_NOMER_PO_PORADKY).value)==str and
                    sheet_0.cell(row=inrow,column=column_pole_NOMER_PO_PORADKY).value.replace(' ','')=='1')):
                    
                    
                    
                    
                    for incol in range(column_pole_NOMER_PO_PORADKY,kol_column):
                        if (sheet_0.cell(row=inrow,column=incol).value==3 or
                            (type(sheet_0.cell(row=inrow,column=incol).value)==str and
                            sheet_0.cell(row=inrow,column=incol).value.replace(' ','')=='3')):
                            
                            nomer_stolbca_strok_kod_is_TN=incol
                            for row_kod in rows_list_to_find:
                                spisok_kodov_is_TN.append(sheet_0.cell(row=row_kod,column=nomer_stolbca_strok_kod_is_TN).value)
                            succes=1
                            raise Exception
        except:
            pass

        if succes!=1:
            messagebox.showerror("Ошибка","Ошибка, в файле "+str(list_of_tovarni_nakladniyi_xl[l])+" не были найдены данные в столбце код (3)")
                                
                        
            
            
                    
        
        
        dva_list_of_kol_and_massa=[]
        for o in range(len(list_of_kol_and_massa)):
            dva_list_of_kol_and_massa.append(list(list_of_kol_and_massa[o]))

        

        for h in range(len(dva_list_of_kol_and_massa)):
            for j in range(len(dva_list_of_kol_and_massa[h])):
                if dva_list_of_kol_and_massa[h][j]==',':
                    dva_list_of_kol_and_massa[h][j]='.'
        
        
        
        
        for z in range(len(list_of_kol_and_massa)):
            list_of_kol_and_massa[z]=''.join(dva_list_of_kol_and_massa[z])
            
        
        list_of_kol=[]
        list_of_massa=[]
        
        for k in list_of_kol_and_massa:
            for s in range(len(k)):
                if k[s]=='(':
                    #list_of_massa.append(float(k[s:]))
                    list_of_kol.append(float(k[:s]))
                    for s2 in range(s,len(k)):
                        if k[s2]==')':
                            to_append=k[s+1:s2].replace(' ','')
                            list_of_massa.append(float(to_append))

        

        
        # ПОИСК ЗНАЧЕНИЯ С ИТОГОВОЙ СУММОЙ БЕЗ НДС
        Itogovaya_summa_bez_NDS=0

        #Поиск номера строки Поля  "Всего по накладной:"


        for n_row in range(rows_list_to_find[-1],kol_row):
            for n_col in range(2,kol_column):
                if sheet_0.cell(row=n_row,column=n_col).value!=None:
                    if type(sheet_0.cell(row=n_row,column=n_col).value)==str:
                        if sheet_0.cell(row=n_row,column=n_col).value.lower()=="всего по накладной:":
                            
                            Itogovaya_summa_bez_NDS=sheet_0.cell(row=n_row,column=column_stolbca_s_summoy_bez_nds).value
                            break
                            
                            
                    
                    
        #формирование списка для записи в exel
        # 1 счет-фактура №
        # 2 дата
        # 3 наименование из ТН (часть)
        # 4 вес
        # 5 стоимость в евро
        # 6 стоимость в долларах
        # 7 стоимость в рублях
        # 8 страна
        # 9 номер дт
        # 10 курс евро
        # 11 курс доллара
        # 12 кол-во
        # 13 код
        # 14 ТН номер
        # 15 дата
        # 16 счет на оплату
        # 17 КОД ВАЛЮТЫ
        # 18 код товара из СФ
        # 19 код товара из ТН
        # 20
        # 21 Порядковый номер товара из СФ
        # 22 Порядковый номер товара из ТН
        # 23 Итоговая сумма без НДС из ТН
        # 24 Итоговая сумма без НДС из СФ

        for i in range(len(rows_list_to_find)):
            row_list=['-' for ij in range(24)]
            row_list[16]=schet_na_oplaty
            row_list[15]=Data_sostavleniya
            row_list[14]=nomer_documenta
            if currency=='Евро':
                row_list[5]=list_of_prices[i]
                row_list[17]=978
            if currency=='Руб':
                row_list[7]=list_of_prices[i]
                row_list[17]=643
            if currency=='Дол':
                row_list[6]=list_of_prices[i]
                row_list[17]=840
            row_list[4]=list_of_massa[i]
            row_list[12]=list_of_kol[i]
            if len(list_of_naimenovaniy_tovar[i])<50:
                row_list[3]=list_of_naimenovaniy_tovar[i]
            else:
                row_list[3]=list_of_naimenovaniy_tovar[i][:50]
            row_list[19]=spisok_kodov_is_TN[i]
            row_list[22]=i+1
            row_list[23]=round(Itogovaya_summa_bez_NDS,1)

            
                
            
            
            print(row_list)
            spisok_massivov_dla_zapici.append(row_list)
        

     
        print()
        print()
        print()
       
    l2marker['bg']="#007dff"
    l2marker['text']='Данные ТН скопированы'


def creating_EXEL_and_copying_there():
    
    if 'spisok_massivov_is_SF' in globals() and 'spisok_massivov_dla_zapici' in globals():
        if len(spisok_massivov_is_SF)>0 and len(spisok_massivov_dla_zapici)>0:
            pass
        else:
            messagebox.showinfo("Отсутствие данных", "Данные из СФ и/или ТН не были загружены")
            return    
    else:
        messagebox.showinfo("Отсутствие данных", "Данные из СФ и/или ТН не были загружены")
        return
   
    try:
        os.chdir(e5.get())
    except OSError:
        messagebox.showerror("Ошибка","Ошибка в пути папки. Директории не существует")
        return

    stock_EXEL_fail_workbook=openpyxl.Workbook()
    sheet_of_stock_EXEL = stock_EXEL_fail_workbook[stock_EXEL_fail_workbook.sheetnames[0]]

    sheet_of_stock_EXEL.cell(row = 1, column = 1).value="Счет-фактура №"
    sheet_of_stock_EXEL.cell(row = 1, column = 2).value="ДатаСФ"
    sheet_of_stock_EXEL.cell(row = 1, column = 3).value="Наименование"
    sheet_of_stock_EXEL.cell(row = 1, column = 4).value="Вес"
    sheet_of_stock_EXEL.cell(row = 1, column = 5).value="Стоимость в евро"
    sheet_of_stock_EXEL.cell(row = 1, column = 6).value="Стоимость в долларах"
    sheet_of_stock_EXEL.cell(row = 1, column = 7).value="Стоимость в рублях"
    sheet_of_stock_EXEL.cell(row = 1, column = 8).value="Страна"
    sheet_of_stock_EXEL.cell(row = 1, column = 9).value="Номер дт"
    sheet_of_stock_EXEL.cell(row = 1, column = 10).value="Курс евро"
    sheet_of_stock_EXEL.cell(row = 1, column = 11).value="Курс доллара"
    sheet_of_stock_EXEL.cell(row = 1, column = 12).value="Кол-во"
    sheet_of_stock_EXEL.cell(row = 1, column = 13).value="Код"
    sheet_of_stock_EXEL.cell(row = 1, column = 14).value="Товарная Накладная"# ТН - это товарная накладная????
    sheet_of_stock_EXEL.cell(row = 1, column = 15).value="ДатаТН"
    sheet_of_stock_EXEL.cell(row = 1, column = 16).value="Счет на оплату"
    sheet_of_stock_EXEL.cell(row = 1, column = 17).value="Код Валюты"
    sheet_of_stock_EXEL.cell(row = 1, column = 18).value="СФ Код товара/ работ, услуг"
    sheet_of_stock_EXEL.cell(row = 1, column = 19).value="ТН Код товара/ работ, услуг"

    list_time=time.localtime()
    beta_name="Стоковый_EXEL "+str(list_time[2])+'.'+str(list_time[1])+'.'+str(list_time[0])+'  '+str(list_time[3])+'-'+str(list_time[4])+".xlsx"
    beta_name=list(beta_name)
    name=''
    for i in range(len(beta_name)):
        if beta_name[i]==":":
            beta_name[i]="-"
    for i in beta_name:
        name+=i

        

    #сращивание двух массивов
    indexi_is_TN_s_parami=[]
    indexi_is_SF_s_parami=[]
    
        
    spisok_stolbcov_is_SF_dla_copy_v_TN=[1,2,8,9,13]
    for n in range(len(spisok_massivov_dla_zapici)):
        for m in range(len(spisok_massivov_is_SF)):
            if (spisok_massivov_dla_zapici[n][19]==spisok_massivov_is_SF[m][18] and     # место нахождения пары !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
                spisok_massivov_dla_zapici[n][22]==spisok_massivov_is_SF[m][21] and
                spisok_massivov_dla_zapici[n][23]==spisok_massivov_is_SF[m][24]): # ИСПРАВИТЬ   
                indexi_is_TN_s_parami.append(n)
                indexi_is_SF_s_parami.append(m)
                for cell_SF in spisok_stolbcov_is_SF_dla_copy_v_TN:
                    spisok_massivov_dla_zapici[n][cell_SF]=spisok_massivov_is_SF[m][cell_SF]

    indexi_is_TN_bez_par=[]
    indexi_is_SF_bez_par=[]
    for i in range(len(spisok_massivov_dla_zapici)):
        if i not in indexi_is_TN_s_parami:
            indexi_is_TN_bez_par.append(i)

    for i in range(len(spisok_massivov_is_SF)):
        if i not in indexi_is_SF_s_parami:
             indexi_is_SF_bez_par.append(i)


    print('Строки из ТН без пары ',indexi_is_TN_bez_par)
    
    print('Строки из СФ без пары ',indexi_is_SF_bez_par)
            
    print('Строки из ТН c парами ',indexi_is_TN_s_parami)
    
    print('Строки из СФ с парами ',indexi_is_SF_s_parami)
    stroka_dla_otobrageniya=' Номер ТН   позиция\n'
    if len(indexi_is_TN_bez_par)>0:
        for i in indexi_is_TN_bez_par:
            stroka_dla_otobrageniya=stroka_dla_otobrageniya+str(spisok_massivov_dla_zapici[i][14])+'               '+str(spisok_massivov_dla_zapici[i][22])+'\n'
            
        messagebox.showinfo("Для некоторых записей из ТН не была найдена пара в СФ", stroka_dla_otobrageniya)

    stroka2_dla_otobrageniya=' Номер СФ   позиция\n'
    if len(indexi_is_SF_bez_par)>0:
        for i in indexi_is_SF_bez_par:
            stroka2_dla_otobrageniya=stroka2_dla_otobrageniya+str(spisok_massivov_is_SF[i][1])+'              '+str(spisok_massivov_is_SF[i][21])+'\n'
            
        messagebox.showinfo("Для некоторых записей из CФ не была найдена пара в ТН", stroka2_dla_otobrageniya)
    
                    
                

            







    
    #ЗАПИСЬ ГЛОБАЛЬНОГО МАССИВА ИЗ ТН
    try:
        for a in range(len(spisok_massivov_dla_zapici)):
            #spisok_massivov_dla_zapici[a]
            for d in range(1,20):
                sheet_of_stock_EXEL.cell(row = a+2, column = d).value=spisok_massivov_dla_zapici[a][d]
        if len(indexi_is_SF_bez_par)>0:
            stroka_dla_vpisania=len(spisok_massivov_dla_zapici)+2
            for istr in indexi_is_SF_bez_par:
                for d in range(1,20):
                    sheet_of_stock_EXEL.cell(row = stroka_dla_vpisania, column = d).value=spisok_massivov_is_SF[istr][d]
                stroka_dla_vpisania+=1
                
            
            
               
    except:
        print("Ошибка записи файлов из ТН в EXEL")
        


    stock_EXEL_fail_workbook.save(name)
    
    print("Создание стокового EXEL завершено")
    messagebox.showinfo("Стоковый EXEL создан", "Стоковый EXEL успешно создан.Для корректной работы нужна перезагрузка. Если вам еще нужна программа - просто запустите ее еще раз")
    root.destroy()
    
     

    


b_spravka.config(command=show_spravka)
b1.config(command=askdir)
b2.config(command=zapusk_kopy)



b1promeg.config(command=askdir_promeg_TN)
b2promeg.config(command=promeg_TN_save)

b_2Val.config(command=ask_Val_xl)
b_3Val.config(command=save_from_Val)
b_4Val.config(command=integration_Val_v_TN)


b21.config(command=askdir_of_SF)
b22.config(command=zapusk_kopy_SF)
b21promeg.config(command=askdir_promeg_SF)
b22promeg.config(command=promeg_SF_save)



b5.config(command=askdir_of_stock)
b52.config(command=creating_EXEL_and_copying_there)
root.mainloop()
